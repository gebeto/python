# -*- coding: utf-8 -*-

import re
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from copy import copy
from openpyxl import (
    load_workbook,
    Workbook
)

color_blue = '8EB4E3'
color_green = 'D7E4BD'
color_yellow = 'FCD5B5'
color_grey = '808080'


center_alignment = Alignment(horizontal="center", vertical="center", wrapText=True)


def bool_value(value, yes='yes', no='no'):
	if value == yes:
		return "1:Да"
	if value == no:
		return "0:Нет"
	if yes == '*':
		return "1:Да"
	return ""

def calculate_column_width(col_type):
	if re.match(r'varchar\(([0-9]+)\)', col_type):
		m = re.match(r'varchar\(([0-9]+)\)', col_type)
		res = int(m.group(1))
		return res if res < 128 else 128
	elif re.match('int', col_type):
		return 10
	elif re.match('real', col_type):
		return 10
	return 10


class PricelistColumn(object):
	def __init__(self, col_fill, col_name, col_type, col_title, col_color='FFFFFF', width=10):
		self.col_fill = col_fill
		self.col_name = col_name
		self.col_type = col_type
		self.col_title = col_title
		self.col_color = col_color

		self.cell_prototype = None
		self.col_index = None
		self.width = calculate_column_width(col_type)
		self.max_width = 0

	def init(self, ws, index):
		self.col_index = index
		current_fill = PatternFill("solid", fgColor=self.col_fill)

		cell1 = ws.cell(row=1, column=index)
		cell1.value = self.col_name
		cell1.fill = current_fill
		cell1.alignment = center_alignment
		# ws.row_dimensions[cell1.column_letter].height = long(100)
		ws.row_dimensions[1].height = 40
		ws.column_dimensions[cell1.column_letter].fill = current_fill

		self.cell_prototype = copy(cell1)
		ws.column_dimensions[cell1.column_letter].width = self.width

		cell2 = ws.cell(row=2, column=index, value=self.col_type)
		cell2.value = self.col_type
		cell2.fill = current_fill
		cell2.alignment = center_alignment
		ws.row_dimensions[2].height = 40

		cell3 = ws.cell(row=3, column=index, value=self.col_title)
		cell3.value = self.col_title
		cell3.fill = PatternFill("solid", fgColor="808080")
		cell3.font = Font(color="FFFFFF")
		cell3.alignment = center_alignment
		ws.row_dimensions[3].height = 60


class BasePricelist(object):
	def __init__(self, title):
		self.title = title
		self.wb = Workbook()
		self.ws = self.wb.active
		self.ws.title = self.title
		self.ws.page_setup.fitToHeight = 1
		self.ws.page_setup.fitToWidth = 1
		self.current_row = 4
		self._initialized = False

		self.columns_dict = {}
		self.columns = [
			PricelistColumn(color_blue,   'Code', 'int', 'Code', width=10),
			PricelistColumn(color_green,  'ModelNameRu', 'varchar(64)', 'Название модели', width=36),
			PricelistColumn(color_green,  'ModelNameUk', 'varchar(64)', 'Название модели (Uk)', width=30),
			PricelistColumn(color_green,  'ModelGroupNameRu', 'varchar(32)', 'Название модельного ряда', width=30),
			PricelistColumn(color_green,  'ModelGroupNameUk', 'varchar(32)', 'Название модельного ряда (Uk)', width=30),
			PricelistColumn(color_green,  'ManufacturerId', 'int', 'Производитель', width=10),
			PricelistColumn(color_green,  'TradeMarkId', 'int', 'Торговая марка', width=10),
			PricelistColumn(color_green,  'ManufacturerCountryId', 'int', 'Страна-производитель', width=10),
			PricelistColumn(color_green,  'VendorCode', 'varchar(16)', 'Код производителя'),
			PricelistColumn(color_yellow, 'Price', 'real', 'Цена у нас на сайте, грн'),
			PricelistColumn(color_yellow, 'PriceUSD', 'real', 'Цена у нас на сайте, $'),
			PricelistColumn(color_yellow, 'PriceRecommendedManufacturer', 'real', 'РРЦ, грн'),
			PricelistColumn(color_yellow, 'PriceRecommendedManufacturerUSD', 'real', 'РРЦ, $'),
			PricelistColumn(color_yellow, 'WholesaleDiscount', 'real', 'Маржа (Оптовая скидка), %'),
			PricelistColumn(color_yellow, 'WholesalePrice', 'real', 'Оптовая цена, грн'),
			PricelistColumn(color_yellow, 'WholesalePriceUSD', 'real', 'Оптовая цена, $'),
			PricelistColumn(color_yellow, 'Available', 'boolean', 'Доступен для заказа у поставщика'),
			PricelistColumn(color_yellow, 'OutOfProduction', 'boolean', 'Снят с производства'),
			PricelistColumn(color_yellow, 'Preorder', 'boolean', 'Доступен только под заказ'),
			PricelistColumn(color_green,  'Recommended', 'boolean', ''),
			PricelistColumn(color_blue,   'IsRelease', 'boolean', 'Показывать на сайте'),
		]

	def save(self, path = None):
		file_path = path if path else "{}.xlsx".format(self.title)
		self.wb.save(file_path)

	def _init_columns(self):
		wb = self.wb
		ws = self.ws

		for column_index, column in enumerate(self.columns):
			self.columns_dict[column.col_name] = column
			column.init(ws, column_index + 1)

		self._initialized = True
		return wb, ws

	def add_row(self, *args):
		if not self._initialized:
			self._init_columns()
		for column_index, item in enumerate(args):
			if type(item) == str:
				column = self.columns[column_index]
				item = item
			elif type(item) in [list, set, tuple]:
				if not item[0] in self.columns_dict: continue
				column = self.columns_dict[item[0]]
				item = item[1]
			cell = self.ws.cell(row=self.current_row, column=column.col_index)
			cell._style = copy(column.cell_prototype._style)
			cell.value = item
			if type(item) != unicode and len(str(item)) < 10:
				cell.alignment = center_alignment
			else:
				cell.alignment = Alignment(horizontal="left", wrapText=True)
		self.current_row += 1




# pricelist = BasePricelist("Hell")
# pricelist.init_columns()
# pricelist.add_row(["Code", "1"], ["ModelNameRu", "Hello"], ["Price", "123"], ["PriceUSD", "1234"])
# pricelist.add_row(["Code", "2"], ["ModelNameRu", "World"], ["Price", "123"], ["PriceUSD", "1234"])
# pricelist.save("/Users/gebeto/Desktop/test.xlsx")

